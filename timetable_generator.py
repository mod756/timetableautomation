import pandas as pd
import random
from datetime import datetime, time, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# Constants
DAYS = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
START_TIME = time(9, 0)
END_TIME = time(18, 30)
TIME_SLOT_DURATION = 30  # minutes (REQ-02)
LECTURE_DURATION = 3    # 1.5 hours = 3 slots
TUTORIAL_DURATION = 2   # 1 hour = 2 slots
LAB_DURATION = 4       # 2 hours = 4 slots
MORNING_BREAK = (time(10, 30), time(11, 0))  # 30 min
LUNCH_BREAK = (time(13, 0), time(14, 0))     # 1 hour
AFTERNOON_BREAK = (time(16, 30), time(17, 0))  # 30 min

def generate_time_slots():
    slots = []
    current_time = datetime.combine(datetime.today(), START_TIME)
    end_time = datetime.combine(datetime.today(), END_TIME)
    
    while current_time < end_time:
        current = current_time.time()
        next_time = current_time + timedelta(minutes=TIME_SLOT_DURATION)
        slots.append((current, next_time.time()))
        current_time = next_time
    return slots

def is_break_time(slot):
    start, _ = slot
    morning_break = MORNING_BREAK[0] <= start < MORNING_BREAK[1]
    lunch_break = LUNCH_BREAK[0] <= start < LUNCH_BREAK[1]
    afternoon_break = AFTERNOON_BREAK[0] <= start < AFTERNOON_BREAK[1]
    return morning_break or lunch_break or afternoon_break

# Load data from three CSV files (REQ-02)
try:
    courses_df = pd.read_csv('courses.csv')
    electives_df = pd.read_csv('electives.csv')
    rooms_df = pd.read_csv('rooms.csv')
    print("Courses columns:", courses_df.columns.tolist())
    print("Electives columns:", electives_df.columns.tolist())
    print("Rooms columns:", rooms_df.columns.tolist())
except FileNotFoundError as e:
    print(f"Error: File {e.filename} not found")
    exit()

# Create faculty name mapping for electives
faculty_map = {}
for _, row in electives_df.iterrows():
    faculty_ids = str(row['faculty_id']).split(';')
    faculty_names = str(row['faculty_name']).split(';')
    for fid, fname in zip(faculty_ids, faculty_names):
        faculty_map[fid] = fname.strip()

def get_faculty_name(faculty_ids, course_code):
    if pd.isna(faculty_ids):
        return "Unknown"
    fids = str(faculty_ids).split(';')
    names = [faculty_map.get(fid, f"Faculty_{fid}") for fid in fids]
    return ', '.join(names)

def assign_room(course_capacity, is_lab, scheduled_rooms, day, start_slot, duration):
    # Prioritize larger rooms for high-capacity courses
    if course_capacity >= 75:
        available_rooms = rooms_df[rooms_df['capacity'] >= 120]  # C002, C003, C004
    else:
        available_rooms = rooms_df[rooms_df['capacity'] >= course_capacity]
    
    # Filter by room type (REQ-03, REQ-08)
    if is_lab:
        available_rooms = available_rooms[available_rooms['room type'].isin(['COMPUTER_LAB', 'HARDWARE_LAB'])]
    else:
        available_rooms = available_rooms[available_rooms['room type'].isin(['LECTURE_ROOM', 'SEATER_120', 'SEATER_240'])]
    
    if available_rooms.empty:
        print(f"No suitable room found for capacity {course_capacity}, lab={is_lab}")
        return None, None
    
    # Check for room conflicts with detailed logging
    for _, room in available_rooms.sample(frac=1).iterrows():  # Randomize room selection
        room_id = room['id']
        room_no = room['room no']
        if room_id not in scheduled_rooms:
            scheduled_rooms[room_id] = {d: set() for d in range(len(DAYS))}
        
        conflict = False
        for i in range(duration):
            if start_slot + i in scheduled_rooms[room_id][day]:
                print(f"Conflict in room {room_no} (ID: {room_id}) for capacity {course_capacity}, day {day}, slot {start_slot+i}")
                conflict = True
                break
        if not conflict:
            print(f"Assigned room {room_no} (ID: {room_id}) for capacity {course_capacity}, day {day}, slot {start_slot}")
            return room_id, room_no
    
    print(f"No available room for capacity {course_capacity}, day {day}, slot {start_slot} after checking all options")
    return None, None

def generate_all_timetables():
    TIME_SLOTS = generate_time_slots()
    wb = Workbook()
    wb.remove(wb.active)
    
    professor_schedule = {}
    classroom_schedule = {}
    max_attempts = 15000  # Increased attempts to ensure scheduling success
    
    # Process each department and semester
    for department in courses_df['DEPARTMENT'].unique():
        for semester in courses_df[courses_df['DEPARTMENT'] == department]['SEMESTER'].unique():
            courses = courses_df[(courses_df['DEPARTMENT'] == department) & 
                               (courses_df['SEMESTER'] == semester)].copy()
            
            if courses.empty:
                continue
            
            # Split into sections if capacity exceeds max room (REQ-03)
            max_room_capacity = rooms_df['capacity'].max()
            sections = {}
            for _, course in courses.iterrows():
                course_id = course['COURSE_ID']
                capacity = course['CAPACITY']
                if capacity > max_room_capacity:
                    sections[course_id] = ['A', 'B']
                else:
                    sections[course_id] = ['A']
            
            # Identify common electives across specific sections (CSE 2A/2B, CSE 6A/6B)
            elective_schedules = {}  # {course_code_type: (day, start_slot, faculty_ids)}
            elective_courses = courses[courses['COMBINED'] == True]
            
            # Create timetables for each section
            section_timetables = {}
            for section in set(sum(sections.values(), [])):
                section_timetables[section] = {
                    'timetable': {day: {slot: {'type': None, 'code': '', 'name': '', 'faculty': '', 'classroom': ''} 
                                for slot in range(len(TIME_SLOTS))} for day in range(len(DAYS))},
                    'course_days': {}
                }
            
            # Schedule common electives first
            elective_baskets = elective_courses['COURSE_CODE'].unique()
            for basket in elective_baskets:
                basket_courses = courses[courses['COURSE_CODE'] == basket]
                for _, course in basket_courses.iterrows():
                    course_id = course['COURSE_ID']
                    code = str(course['COURSE_CODE'])
                    name = str(course['COURSE_NAME'])
                    faculty_ids = str(course['FACULTY_ID'])
                    faculty = get_faculty_name(faculty_ids, code)
                    capacity = course['CAPACITY']
                    l = int(course['L']) if pd.notna(course['L']) else 0
                    t = int(course['T']) if pd.notna(course['T']) else 0
                    p = int(course['P']) if pd.notna(course['P']) else 0
                    
                    # Determine sections for this course
                    course_sections = sections[course_id]
                    if not course_sections:
                        continue
                    
                    # Identify specific section pairs for synchronization
                    target_sections = []
                    if department == 'CSE' and semester == 2 and 'A' in course_sections and 'B' in course_sections:
                        target_sections = ['A', 'B']  # CSE 2A and 2B
                    elif department == 'CSE' and semester == 6 and 'A' in course_sections and 'B' in course_sections:
                        target_sections = ['A', 'B']  # CSE 6A and 6B
                    
                    if target_sections:
                        # Schedule once for the target sections
                        if code not in [k.split('_')[0] for k in elective_schedules.keys()]:
                            # Schedule lectures
                            for _ in range(l):
                                scheduled = False
                                attempts = 0
                                while not scheduled and attempts < max_attempts:
                                    day = random.randint(0, len(DAYS)-1)
                                    start_slot = random.randint(0, len(TIME_SLOTS)-LECTURE_DURATION)
                                    slots_free = True
                                    for i in range(LECTURE_DURATION):
                                        if is_break_time(TIME_SLOTS[start_slot+i]):
                                            slots_free = False
                                            break
                                        for section in target_sections:
                                            if (section_timetables[section]['timetable'][day][start_slot+i]['type'] is not None):
                                                slots_free = False
                                                break
                                    if not slots_free:
                                        attempts += 1
                                        continue
                                    
                                    section_rooms = {}
                                    all_rooms_assigned = True
                                    for section in target_sections:
                                        section_capacity = capacity // len(target_sections)
                                        room_id, classroom = assign_room(section_capacity, False, classroom_schedule, day, start_slot, LECTURE_DURATION)
                                        if room_id is None:
                                            all_rooms_assigned = False
                                            break
                                        section_rooms[section] = (room_id, classroom)
                                    
                                    if all_rooms_assigned:
                                        slots_free = True
                                        for fid in faculty_ids.split(';'):
                                            if fid not in professor_schedule:
                                                professor_schedule[fid] = {d: set() for d in range(len(DAYS))}
                                            for i in range(LECTURE_DURATION):
                                                if start_slot+i in professor_schedule[fid][day]:
                                                    slots_free = False
                                                    break
                                        if slots_free:
                                            print(f"Scheduling lecture {code} on {DAYS[day]} at {TIME_SLOTS[start_slot][0]} for sections {target_sections}")
                                            elective_schedules[f"{code}_LEC"] = (day, start_slot, faculty_ids)
                                            for section in target_sections:
                                                timetable = section_timetables[section]['timetable']
                                                room_id, classroom = section_rooms[section]
                                                for i in range(LECTURE_DURATION):
                                                    timetable[day][start_slot+i]['type'] = 'LEC'
                                                    timetable[day][start_slot+i]['code'] = code if i == 0 else ''
                                                    timetable[day][start_slot+i]['name'] = name if i == 0 else ''
                                                    timetable[day][start_slot+i]['faculty'] = faculty if i == 0 else ''
                                                    timetable[day][start_slot+i]['classroom'] = classroom if i == 0 else ''
                                                    classroom_schedule[room_id][day].add(start_slot+i)
                                                section_timetables[section]['course_days'].setdefault(course_id, set()).add(day)
                                            for fid in faculty_ids.split(';'):
                                                for i in range(LECTURE_DURATION):
                                                    professor_schedule[fid][day].add(start_slot+i)
                                            scheduled = True
                                    attempts += 1
                                    if attempts >= max_attempts:
                                        print(f"Failed to schedule lecture {code} after {max_attempts} attempts due to room/faculty conflicts")
                            
                            # Schedule tutorials
                            for _ in range(t):
                                scheduled = False
                                attempts = 0
                                while not scheduled and attempts < max_attempts:
                                    day = random.randint(0, len(DAYS)-1)
                                    start_slot = random.randint(0, len(TIME_SLOTS)-TUTORIAL_DURATION)
                                    slots_free = True
                                    max_tut_days = 3  # Allow up to 3 days for tutorials
                                    if course_id in section_timetables[section]['course_days'] and len(section_timetables[section]['course_days'][course_id]) >= max_tut_days:
                                        break  # Skip if max tutorial days reached
                                    for i in range(TUTORIAL_DURATION):
                                        if is_break_time(TIME_SLOTS[start_slot+i]):
                                            slots_free = False
                                            break
                                        if timetable[day][start_slot+i]['type'] is not None:
                                            slots_free = False
                                            break
                                    if not slots_free:
                                        attempts += 1
                                        continue
                                    
                                    room_id, classroom = assign_room(capacity // len(course_sections), False, classroom_schedule, day, start_slot, TUTORIAL_DURATION)
                                    if room_id is None:
                                        attempts += 1
                                        continue
                                    
                                    slots_free = True
                                    for fid in faculty_ids.split(';'):
                                        if fid not in professor_schedule:
                                            professor_schedule[fid] = {d: set() for d in range(len(DAYS))}
                                        for i in range(TUTORIAL_DURATION):
                                            if start_slot+i in professor_schedule[fid][day]:
                                                slots_free = False
                                                break
                                    if slots_free:
                                        print(f"Scheduling tutorial {code} on {DAYS[day]} at {TIME_SLOTS[start_slot][0]} for section {section}")
                                        for i in range(TUTORIAL_DURATION):
                                            timetable[day][start_slot+i]['type'] = 'TUT'
                                            timetable[day][start_slot+i]['code'] = code if i == 0 else ''
                                            timetable[day][start_slot+i]['name'] = name if i == 0 else ''
                                            timetable[day][start_slot+i]['faculty'] = faculty if i == 0 else ''
                                            timetable[day][start_slot+i]['classroom'] = classroom if i == 0 else ''
                                            classroom_schedule[room_id][day].add(start_slot+i)
                                        section_timetables[section]['course_days'].setdefault(course_id, set()).add(day)
                                        for fid in faculty_ids.split(';'):
                                            for i in range(TUTORIAL_DURATION):
                                                professor_schedule[fid][day].add(start_slot+i)
                                        scheduled = True
                                    attempts += 1
                                    if attempts >= max_attempts:
                                        print(f"Failed to schedule tutorial {code} for section {section} after {max_attempts} attempts due to room/faculty conflicts")
                            
                            # Schedule labs with strength division
                            if p > 0:
                                batch_size = capacity // 2
                                scheduled = False
                                attempts = 0
                                while not scheduled and attempts < max_attempts:
                                    day = random.randint(0, len(DAYS)-1)
                                    start_slot = random.randint(0, len(TIME_SLOTS)-LAB_DURATION)
                                    slots_free = True
                                    for i in range(LAB_DURATION):
                                        if is_break_time(TIME_SLOTS[start_slot+i]):
                                            slots_free = False
                                            break
                                        for section in target_sections:
                                            if (section_timetables[section]['timetable'][day][start_slot+i]['type'] is not None):
                                                slots_free = False
                                                break
                                    if not slots_free:
                                        attempts += 1
                                        continue
                                    
                                    room1_id, room1_no = assign_room(batch_size, True, classroom_schedule, day, start_slot, LAB_DURATION)
                                    room2_id, room2_no = assign_room(batch_size, True, classroom_schedule, day, start_slot, LAB_DURATION)
                                    if room1_id is None or room2_id is None or room1_id == room2_id:
                                        attempts += 1
                                        continue
                                    
                                    slots_free = True
                                    for fid in faculty_ids.split(';'):
                                        if fid not in professor_schedule:
                                            professor_schedule[fid] = {d: set() for d in range(len(DAYS))}
                                        for i in range(LAB_DURATION):
                                            if start_slot+i in professor_schedule[fid][day]:
                                                slots_free = False
                                                break
                                    if not slots_free:
                                        attempts += 1
                                        continue
                                    
                                    print(f"Scheduling lab {code} on {DAYS[day]} at {TIME_SLOTS[start_slot][0]} with rooms {room1_no} and {room2_no} for sections {target_sections}")
                                    elective_schedules[f"{code}_LAB"] = (day, start_slot, faculty_ids)
                                    for section in target_sections:
                                        timetable = section_timetables[section]['timetable']
                                        room_id = room1_id if section == target_sections[0] else room2_id
                                        classroom = room1_no if section == target_sections[0] else room2_no
                                        for i in range(LAB_DURATION):
                                            timetable[day][start_slot+i]['type'] = 'LAB'
                                            timetable[day][start_slot+i]['code'] = code if i == 0 else ''
                                            timetable[day][start_slot+i]['name'] = name if i == 0 else ''
                                            timetable[day][start_slot+i]['faculty'] = faculty if i == 0 else ''
                                            timetable[day][start_slot+i]['classroom'] = classroom if i == 0 else ''
                                            classroom_schedule[room_id][day].add(start_slot+i)
                                        section_timetables[section]['course_days'].setdefault(course_id, set()).add(day)
                                    for fid in faculty_ids.split(';'):
                                        for i in range(LAB_DURATION):
                                            professor_schedule[fid][day].add(start_slot+i)
                                    scheduled = True
                                    attempts += 1
                                    if attempts >= max_attempts:
                                        print(f"Failed to schedule lab {code} after {max_attempts} attempts due to room/faculty conflicts")
            
            # Schedule core courses for each section
            core_courses = courses[courses['COMBINED'] == False]
            for section in section_timetables:
                timetable = section_timetables[section]['timetable']
                course_days = section_timetables[section]['course_days']
                
                for _, course in core_courses.iterrows():
                    course_id = course['COURSE_ID']
                    code = str(course['COURSE_CODE'])
                    name = str(course['COURSE_NAME'])
                    faculty_ids = str(course['FACULTY_ID'])
                    faculty = get_faculty_name(faculty_ids, code)
                    capacity = course['CAPACITY']
                    l = int(course['L']) if pd.notna(course['L']) else 0
                    t = int(course['T']) if pd.notna(course['T']) else 0
                    p = int(course['P']) if pd.notna(course['P']) else 0
                    
                    if section not in sections[course_id]:
                        continue
                    
                    course_days.setdefault(course_id, set())
                    max_days = 2  # Limit to 2 days for lectures/labs
                    
                    # Schedule labs with strength division
                    if p > 0:
                        batch_size = capacity // 2
                        scheduled = False
                        attempts = 0
                        while not scheduled and attempts < max_attempts:
                            day = random.randint(0, len(DAYS)-1)
                            if len(TIME_SLOTS) >= LAB_DURATION and len(course_days.get(course_id, set())) < max_days:
                                start_slot = random.randint(0, len(TIME_SLOTS)-LAB_DURATION)
                                if course_id in course_days and day in course_days[course_id]:
                                    continue  # Avoid same day if max days reached
                                
                                slots_free = True
                                for i in range(LAB_DURATION):
                                    if (is_break_time(TIME_SLOTS[start_slot+i]) or
                                        timetable[day][start_slot+i]['type'] is not None):
                                        slots_free = False
                                        break
                                
                                room1_id, room1_no = assign_room(batch_size, True, classroom_schedule, day, start_slot, LAB_DURATION)
                                room2_id, room2_no = assign_room(batch_size, True, classroom_schedule, day, start_slot, LAB_DURATION)
                                if room1_id is None or room2_id is None or room1_id == room2_id:
                                    attempts += 1
                                    continue
                                
                                for fid in faculty_ids.split(';'):
                                    if fid not in professor_schedule:
                                        professor_schedule[fid] = {d: set() for d in range(len(DAYS))}
                                    for i in range(LAB_DURATION):
                                        if start_slot+i in professor_schedule[fid][day]:
                                            slots_free = False
                                            break
                                
                                if slots_free:
                                    print(f"Scheduling lab {code} (section {section}) on {DAYS[day]} at {TIME_SLOTS[start_slot][0]} with rooms {room1_no} and {room2_no}")
                                    room_id = room1_id
                                    classroom = room1_no
                                    for i in range(LAB_DURATION):
                                        timetable[day][start_slot+i]['type'] = 'LAB'
                                        timetable[day][start_slot+i]['code'] = code if i == 0 else ''
                                        timetable[day][start_slot+i]['name'] = name if i == 0 else ''
                                        timetable[day][start_slot+i]['faculty'] = faculty if i == 0 else ''
                                        timetable[day][start_slot+i]['classroom'] = classroom if i == 0 else ''
                                        classroom_schedule[room_id][day].add(start_slot+i)
                                    # Simulate second room (for visualization)
                                    room_id = room2_id
                                    classroom = room2_no
                                    for i in range(LAB_DURATION):
                                        timetable[day][start_slot+i]['classroom'] = f"{timetable[day][start_slot+i]['classroom']}/{classroom}"
                                        classroom_schedule[room_id][day].add(start_slot+i)
                                    course_days[course_id].add(day)
                                    for fid in faculty_ids.split(';'):
                                        for i in range(LAB_DURATION):
                                            professor_schedule[fid][day].add(start_slot+i)
                                    scheduled = True
                            attempts += 1
                            if attempts >= max_attempts:
                                print(f"Failed to schedule lab {code} (section {section}) after {max_attempts} attempts due to room/faculty conflicts")
                    
                    # Schedule lectures
                    for _ in range(l):
                        scheduled = False
                        attempts = 0
                        while not scheduled and attempts < max_attempts:
                            day = random.randint(0, len(DAYS)-1)
                            if len(TIME_SLOTS) >= LECTURE_DURATION and len(course_days.get(course_id, set())) < max_days:
                                start_slot = random.randint(0, len(TIME_SLOTS)-LECTURE_DURATION)
                                if course_id in course_days and day in course_days[course_id]:
                                    continue  # Avoid same day if max days reached
                                
                                slots_free = True
                                for i in range(LECTURE_DURATION):
                                    if (is_break_time(TIME_SLOTS[start_slot+i]) or
                                        timetable[day][start_slot+i]['type'] is not None):
                                        slots_free = False
                                        break
                                
                                room_id, classroom = assign_room(capacity, False, classroom_schedule, day, start_slot, LECTURE_DURATION)
                                if room_id is None:
                                    attempts += 1
                                    continue
                                
                                for fid in faculty_ids.split(';'):
                                    if fid not in professor_schedule:
                                        professor_schedule[fid] = {d: set() for d in range(len(DAYS))}
                                    for i in range(LECTURE_DURATION):
                                        if start_slot+i in professor_schedule[fid][day]:
                                            slots_free = False
                                            break
                                
                                if slots_free:
                                    print(f"Scheduling lecture {code} (section {section}) on {DAYS[day]} at {TIME_SLOTS[start_slot][0]}")
                                    for i in range(LECTURE_DURATION):
                                        timetable[day][start_slot+i]['type'] = 'LEC'
                                        timetable[day][start_slot+i]['code'] = code if i == 0 else ''
                                        timetable[day][start_slot+i]['name'] = name if i == 0 else ''
                                        timetable[day][start_slot+i]['faculty'] = faculty if i == 0 else ''
                                        timetable[day][start_slot+i]['classroom'] = classroom if i == 0 else ''
                                        for fid in faculty_ids.split(';'):
                                            professor_schedule[fid][day].add(start_slot+i)
                                        classroom_schedule[room_id][day].add(start_slot+i)
                                    course_days[course_id].add(day)
                                    scheduled = True
                            attempts += 1
                            if attempts >= max_attempts:
                                print(f"Failed to schedule lecture {code} (section {section}) after {max_attempts} attempts due to room/faculty conflicts")
                    
                    # Schedule tutorials
                    for _ in range(t):
                        if t > 0:  # Debug: Ensure tutorials are attempted
                            print(f"Attempting to schedule {t} tutorials for {code} (section {section})")
                        scheduled = False
                        attempts = 0
                        while not scheduled and attempts < max_attempts:
                            day = random.randint(0, len(DAYS)-1)
                            if len(TIME_SLOTS) >= TUTORIAL_DURATION:
                                start_slot = random.randint(0, len(TIME_SLOTS)-TUTORIAL_DURATION)
                                max_tut_days = 3  # Allow up to 3 days for tutorials
                                if course_id in course_days and len(course_days[course_id]) >= max_tut_days:
                                    break  # Skip if max tutorial days reached
                                
                                slots_free = True
                                for i in range(TUTORIAL_DURATION):
                                    if (is_break_time(TIME_SLOTS[start_slot+i]) or
                                        timetable[day][start_slot+i]['type'] is not None):
                                        slots_free = False
                                        break
                                
                                room_id, classroom = assign_room(capacity // len(sections[course_id]), False, classroom_schedule, day, start_slot, TUTORIAL_DURATION)
                                if room_id is None:
                                    attempts += 1
                                    continue
                                
                                for fid in faculty_ids.split(';'):
                                    if fid not in professor_schedule:
                                        professor_schedule[fid] = {d: set() for d in range(len(DAYS))}
                                    for i in range(TUTORIAL_DURATION):
                                        if start_slot+i in professor_schedule[fid][day]:
                                            slots_free = False
                                            break
                                
                                if slots_free:
                                    print(f"Scheduling tutorial {code} (section {section}) on {DAYS[day]} at {TIME_SLOTS[start_slot][0]}")
                                    for i in range(TUTORIAL_DURATION):
                                        timetable[day][start_slot+i]['type'] = 'TUT'
                                        timetable[day][start_slot+i]['code'] = code if i == 0 else ''
                                        timetable[day][start_slot+i]['name'] = name if i == 0 else ''
                                        timetable[day][start_slot+i]['faculty'] = faculty if i == 0 else ''
                                        timetable[day][start_slot+i]['classroom'] = classroom if i == 0 else ''
                                        classroom_schedule[room_id][day].add(start_slot+i)
                                    course_days[course_id].add(day)
                                    for fid in faculty_ids.split(';'):
                                        for i in range(TUTORIAL_DURATION):
                                            professor_schedule[fid][day].add(start_slot+i)
                                    scheduled = True
                            attempts += 1
                            if attempts >= max_attempts:
                                print(f"Failed to schedule tutorial {code} (section {section}) after {max_attempts} attempts due to room/faculty conflicts")
            
            # Write timetables to worksheets (REQ-14)
            for section in section_timetables:
                ws = wb.create_sheet(title=f"{department}_{semester}_{section}")
                timetable = section_timetables[section]['timetable']
                
                header = ['Day'] + [f"{slot[0].strftime('%H:%M')}-{slot[1].strftime('%H:%M')}" for slot in TIME_SLOTS]
                ws.append(header)
                
                # Formatting
                header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                lec_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
                lab_fill = PatternFill(start_color="98FB98", end_color="98FB98", fill_type="solid")
                tut_fill = PatternFill(start_color="FFE4E1", end_color="FFE4E1", fill_type="solid")
                break_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                border = Border(left=Side(style='thin'), right=Side(style='thin'),
                              top=Side(style='thin'), bottom=Side(style='thin'))
                header_font = Font(bold=True)
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Fill data
                for day_idx, day in enumerate(DAYS):
                    row_num = day_idx + 2
                    ws.append([day])
                    merge_ranges = []
                    
                    for slot_idx in range(len(TIME_SLOTS)):
                        cell_value = ''
                        cell_fill = None
                        
                        if is_break_time(TIME_SLOTS[slot_idx]):
                            cell_value = "BREAK"
                            cell_fill = break_fill
                        elif timetable[day_idx][slot_idx]['type']:
                            if timetable[day_idx][slot_idx]['code']:
                                activity_type = timetable[day_idx][slot_idx]['type']
                                if activity_type == 'LEC':
                                    duration = LECTURE_DURATION
                                    cell_fill = lec_fill
                                elif activity_type == 'LAB':
                                    duration = LAB_DURATION
                                    cell_fill = lab_fill
                                elif activity_type == 'TUT':
                                    duration = TUTORIAL_DURATION
                                    cell_fill = tut_fill
                                
                                start_col = get_column_letter(slot_idx + 2)
                                end_col = get_column_letter(slot_idx + duration + 1)
                                merge_ranges.append(f"{start_col}{row_num}:{end_col}{row_num}")
                                
                                code = timetable[day_idx][slot_idx]['code']
                                classroom = timetable[day_idx][slot_idx]['classroom']
                                cell_value = f"{code} {activity_type}\n{classroom}"
                        
                        cell = ws.cell(row=row_num, column=slot_idx+2, value=cell_value)
                        if cell_fill:
                            cell.fill = cell_fill
                        cell.border = border
                        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                    
                    for merge_range in merge_ranges:
                        ws.merge_cells(merge_range)
                
                # Adjust dimensions
                for col_idx in range(1, len(TIME_SLOTS)+2):
                    ws.column_dimensions[get_column_letter(col_idx)].width = 15
                for row in ws.iter_rows(min_row=2, max_row=len(DAYS)+1):
                    ws.row_dimensions[row[0].row].height = 40
    
    wb.save("timetables.xlsx")
    print("Timetables saved to timetables.xlsx")

if __name__ == "__main__":
    generate_all_timetables()